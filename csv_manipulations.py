import os
import csv
from math import floor

import numpy as np
from xlsxwriter import Workbook

from tools import Person
import constants as cst
from feedback import rotated_feedback_comparison

def extract_values(path, file, data):
    with open(path + file) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')

        splitted = file.split('_')
        flaw_name = ' '.join(splitted[1:-1])

        # Can't use enumerate since there are empty lines in the csv file
        i = 0
        for row in csv_reader:
            if not row:
                continue

            if flaw_name == 'leaning':
                data[i][0] = row[0]
                data[i][1] = row[1]

            elif flaw_name == 'elbow move':
                data[i][2] = row[0]
                data[i][3] = row[1]

            elif flaw_name == 'javelin':
                data[i][4] = row[0]
                data[i][5] = row[1]

            elif flaw_name == 'align arm':
                data[i][6] = row[0]
                data[i][7] = row[1]

            i += 1

    return data


def write_merged_data(path, name, data):
    workbook = Workbook(f'{path}{name}_merged.xlsx')
    worksheet = workbook.add_worksheet(name)

    for i, row in enumerate(data):
        i_with_linejump = i + floor(i / 9)

        for j, value in enumerate(row):
            worksheet.write(f'{chr(65+j)}{i_with_linejump+1}', value)

    workbook.close()


def merge_same_names_xlsx():

    path = './csv_test/'
    files = os.listdir(path)

    all_names = set()

    for file in files:
        if '.csv' in file:
            all_names.add(file.split('_')[0])

    for name in all_names:
        data = np.zeros((36, 8), dtype='float32')
        for file in files:
            if name in file:
                data = extract_values(path, file, data)

        write_merged_data(path, name, data)


def export_advices_to_xlsx(path):
    from xlsxwriter import Workbook

    expert = Person(r'', 'aurel', 'Right')

    workbook = Workbook('new_advices_after_rotation.xlsx')

    for student in cst.students_list:
        print(f'Processing {student.name}...')
        worksheet = workbook.add_worksheet(student.name)
        worksheet.merge_range('A1:B1', 'New')
        worksheet.merge_range('C1:D1', 'Old')
        for i in range(4):
            advices = rotated_feedback_comparison(expert, student, path, begin=i * 9, end=(i * 9) + 9)
            worksheet.write(f'A{i+2}', advices[0])
            worksheet.write(f'B{i+2}', advices[1])

        print('Done.')

    workbook.close()


def merge_xlsx():
    from openpyxl import load_workbook
    wb_results = load_workbook('csv_test/results_2.xlsx')
    wb_comparison = load_workbook('new_advices_after_rotation.xlsx')

    for student in cst.students_list:
        print(f'Processing {student.name}')
        try:
            sheet_results = wb_results[student.name[:-1]]
        except KeyError:
            sheet_results = wb_results['Saupin']
        sheet_comparison = wb_comparison[student.name]

        for i in range(3):
            result_string = sheet_results[f'E{((i+1)*10 + 3)}'].value
            result_string = result_string.split('/')
            sheet_comparison[f'C{i+2}'].value = result_string[0]
            sheet_comparison[f'D{i+2}'].value = result_string[1]

    wb_comparison.save('new_advices_comparison_one_sheet.xlsx')


def one_sheet_xlsx():
    from openpyxl import Workbook, load_workbook
    wb_results = load_workbook('csv_test/results_2.xlsx')
    wb_comparison = load_workbook('new_advices_after_rotation.xlsx')
    wb_final = Workbook()
    final_sheet = wb_final.active

    final_sheet[f'A1'].value = 'Nom'
    final_sheet[f'B1'].value = 'Old'
    final_sheet[f'D1'].value = 'New'


    for row, student in enumerate(cst.students_list):
        good = "True"

        # writting offset
        wo = 3 * row

        print(f'Processing {student.name}')
        try:
            sheet_results = wb_results[student.name[:-1]]
        except KeyError:
            sheet_results = wb_results['Saupin']
        sheet_comparison = wb_comparison[student.name]


        final_sheet[f'A{1 + 1 + wo}'].value = student.name

        for i in range(3):
            result_string = sheet_results[f'E{((i+1)*10 + 3)}'].value
            result_string = result_string.split('/')
            comparison_string = sheet_comparison[f'A{i+2}:B{i+2}']
            comparison_string = [sheet_comparison[f'A{i+2}'].value, sheet_comparison[f'B{i+2}'].value]

            result_string[0] = result_string[0].replace("_", " ").replace("Me\xa0: ", "").lower().strip()
            result_string[1] = result_string[1].replace("_", " ").replace("Me\xa0: ", "").lower().strip()
            comparison_string[0] = comparison_string[0].replace("_", " ").replace("Me : ", "").lower().strip()
            comparison_string[1] = comparison_string[1].replace("_", " ").replace("Me : ", "").lower().strip()

            final_sheet[f'B{i + 1 + 1 + wo}'].value = result_string[0]
            final_sheet[f'C{i + 1 + 1 + wo}'].value = result_string[1]
            final_sheet[f'D{i + 1 + 1 + wo}'].value = comparison_string[0]
            final_sheet[f'E{i + 1 + 1 + wo}'].value = comparison_string[1]

            if result_string[0] != comparison_string[0] or result_string[1] != comparison_string[1]:
                good = "False"

        final_sheet[f'G{wo + 3}'].value = good


    wb_final.save('blblbl.xlsx')


def main():
    merge_same_names_xlsx()


if __name__ == '__main__':
    main()
