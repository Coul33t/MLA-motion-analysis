import os
import stat
from shutil import rmtree
from distutils.dir_util import copy_tree

from sklearn.decomposition import PCA

from data_import import json_import

from data_processing import (run_clustering,
                             data_gathering,
                             plot_good_vs_student_all_data)

from data_visualization import (plot_PCA,
                                multi_plot_PCA,
                                plot_all_defaults)

from feedback_tools import *

from tools import Person, merge_list_of_dictionnaries

import constants as cst

from constants import problemes_et_solutions as problems_and_advices

def import_data(path, import_find):
    original_data = []

    # Gathering the data
    # If it's a list, then we have to import multiple people's data
    if isinstance(import_find[0], list):
        for to_import in import_find[0]:
            original_data.extend(json_import(path, to_import))
    # Else, it's just a name, so we import their data
    else:
        original_data = json_import(path, import_find)

    if not original_data:
        print('ERROR: no data found (check your names).')
        return

    return original_data

def feedback():
    path = r'C:/Users/quentin/Documents/Programmation/C++/MLA/Data/alldartsdescriptors/students/mixed'
    # Expert Data
    name = 'aurel'
    expert_data = import_data(path, name)
    # Student data
    name = 'me'
    student_data = import_data(path, name)

    # Setting the laterality
    for motion in expert_data:
        motion.laterality = 'Right'

    for motion in student_data:
        motion.laterality = 'Left'

    # List of datatypes and joint to process
    # default to check: {Descriptor: [{joint, laterality or not (check left for lefthanded and vice versa)},
    #                                  other joint, laterality or not}],
    #                    Other descriptor: [{joint, laterality or not}]
    #                   }
    datatype_joints_list = []

    datatype_joints_list.append(['leaning', {'MeanSpeed': [{'joint': 'LeftShoulder', 'laterality': False},
                                                           {'joint': 'RightShoulder', 'laterality': False}]
                                }])

    datatype_joints_list.append(['elbow_move', {'MeanSpeed': [{'joint': 'LeftArm', 'laterality': True},
                                                              {'joint': 'LeftShoulder', 'laterality': True}]
                                }])

    datatype_joints_list.append(['javelin', {'PosX': [{'joint': 'LeftHand', 'laterality': True},
                                                      {'joint': 'Head',     'laterality': False}],
                                             'PosY': [{'joint': 'LeftHand', 'laterality': True},
                                                      {'joint': 'Head',     'laterality': False}],
                                             'PosZ': [{'joint': 'LeftHand', 'laterality': True},
                                                      {'joint': 'Head',     'laterality': False}]
                                            }])

    datatype_joints_list.append(['align_arm', {'BoundingBoxMinusX': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxPlusX':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxMinusY': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxPlusY':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxMinusZ': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxPlusZ':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}]
                                              }])



    # Scaling and normalisaing (nor not) the data
    # Useful for DBSCAN for example
    scale = False
    normalise = False

    # Setting the data repartition
    expert_data_repartion = {'good': [x+1 for x in range(10)],
                     'leaning': [x+1 for x in range(10, 20)],
                     'javelin': [x+1 for x in range(20, 30)],
                     'align_arm': [x+1 for x in range(30, 40)],
                     'elbow_move': [x+1 for x in range(40, 50)]}

    # Algorithm to test
    algos = {'k-means': {'n_clusters': 2}}

    distances_and_clusters = []
    results = []
    std_features_all = []

    for problem in datatype_joints_list:

        datatype_joints = problem[1]
        expert_sub_data = expert_data[:10] + expert_data[min(expert_data_repartion[problem[0]])-1:max(expert_data_repartion[problem[0]])]

        for algo, param in algos.items():
            print(problem[0])

            model = run_clustering(path, expert_sub_data, name, validate_data=False,
                                   datatype_joints=datatype_joints, algorithm=algo,
                                   parameters=param, scale_features=scale, normalise_features=normalise,
                                   true_labels=None, verbose=False, to_file=True, to_json=True, return_data=True)

        # Taking the student features
        std_features = data_gathering(student_data, datatype_joints)

        # If the expert data has been scaled, do the same for the student's ones
        if scale:
            from sklearn.preprocessing import RobustScaler
            scaler = RobustScaler()
            std_features = scaler.fit_transform(std_features)
        # Same for normalising
        if normalise:
            from sklearn.preprocessing import MinMaxScaler
            min_max_scaler = MinMaxScaler()
            std_features = min_max_scaler.fit_transform(std_features)

        # Compute the centroid of the student's features (Euclidean distance for now)
        std_centroid = get_centroid(std_features)

        # For the rest of the algorithm, if there are more than 2 features,
        # we run the data through a PCA for the next steps
        if len(model[0].cluster_centers_[0]) > 2:
            pca = PCA(n_components=2, copy=True)
            pca.fit(model[2])

            model[2] = pca.transform(model[2])
            model[0].cluster_centers_ = pca.transform(model[0].cluster_centers_)
            std_centroid = pca.transform(std_centroid.reshape(1, -1))[0]
            std_features = pca.transform(std_features)

        # Compute the distance from the student's centroid to the expert's ones
        distances_to_centroid = compute_distance(model[0].cluster_centers_, std_centroid)

        # Get the distance from the student's centroid to the line between the two expert's centroids
        distance_from_line = get_distance_from_expert_centoids_line(model[0].cluster_centers_, std_centroid)
        # Used to check if the student's centroid is between the expert centroids (diamond shape)
        distance_from_line /= (dst_pts(model[0].cluster_centers_[0], model[0].cluster_centers_[1]) / 2)

        # Normalise the distances to the centroids
        summ = sum(distances_to_centroid)
        for i, distance in enumerate(distances_to_centroid):
            distances_to_centroid[i] = distance/summ

        # Get the most probable cluster label for expert data
        clusters_label = get_cluster_label(expert_sub_data, expert_data_repartion, model[0].labels_)
        # Display the closeness of the student's data to each expert cluster
        mix_res = mix(distances_to_centroid, clusters_label, distance_from_line)
        distances_and_clusters.append(mix_res)

        res = (model[0], model[2], model[0].labels_, model[1], problem[0], std_features, get_trapezoid(model, std_centroid), get_circle(model, std_centroid))
        results.append(res)
        std_features_all.append(std_features)

    models = [x[0] for x in results]
    features = [x[1] for x in results]

    for i, feat in enumerate(features):
        features[i] = np.concatenate([feat, std_features_all[i]], axis=0)

    centroids = [x[0].cluster_centers_ for x in results]
    labels = [x[2] for x in results]

    for i, lab in enumerate(labels):
        max_val = max(lab)
        labels[i] = np.concatenate([lab, [max_val+1 for x in range(len(std_features))]], axis=0)

    trapezoids = [x[6] for x in results]
    circles = [x[7] for x in results]

    sss = [x[3]['ss'] for x in results]
    names = ['k-means ' + x[4] for x in results]
    title = 'None'

    clusters_names = [x[0] for x in distances_and_clusters]

    std_data = [[std_feat, [max(labels[i])+1 for x in range(len(std_features_all[i]))]] for i, std_feat in enumerate(std_features_all)]

    multi_plot_PCA(features, labels, clusters_names, names, models, sss, title, trapezoids, circles, only_centroids=True, centroids=centroids, std_data=std_data)

def only_feedback(expert, student, path):
    folders_path = path
    if folders_path.split('/')[-1] == 'mixed':
        folders_path = "/".join(path.split('/')[:-1])

    if not take_last_data(folders_path, student, expert, number=9):
        print(f'ERROR: {folders_path} does not exists')
        return

    # Expert Data
    expert_data = import_data(path, expert.name)
    # Student data
    student_data = import_data(path, student.name)

    # Setting the laterality
    for motion in expert_data:
        motion.laterality = expert.laterality

    for motion in student_data:
        motion.laterality = student.laterality

    # List of datatypes and joint to process
    # default to check: {Descriptor: [{joint, laterality or not (check left for lefthanded and vice versa)},
    #                                  other joint, laterality or not}],
    #                    Other descriptor: [{joint, laterality or not}]
    #                   }
    datatype_joints_list = []

    datatype_joints_list.append(['leaning', {'MeanSpeed': [{'joint': 'LeftShoulder', 'laterality': False},
                                                           {'joint': 'RightShoulder', 'laterality': False}]
                                }])

    datatype_joints_list.append(['elbow_move', {'MeanSpeed': [{'joint': 'LeftArm', 'laterality': True},
                                                              {'joint': 'LeftShoulder', 'laterality': True}]
                                }])

    datatype_joints_list.append(['javelin', {'PosX': [{'joint': 'LeftHand', 'laterality': True},
                                                      {'joint': 'Head',     'laterality': False}],
                                             'PosY': [{'joint': 'LeftHand', 'laterality': True},
                                                      {'joint': 'Head',     'laterality': False}],
                                             'PosZ': [{'joint': 'LeftHand', 'laterality': True},
                                                      {'joint': 'Head',     'laterality': False}]
                                            }])

    datatype_joints_list.append(['align_arm', {'BoundingBoxMinusX': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxPlusX':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxMinusY': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxPlusY':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxMinusZ': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxPlusZ':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}]
                                              }])


    # Scaling and normalisaing (nor not) the data
    # Useful for DBSCAN for example
    scale = False
    normalise = False

    expert_data_repartion = {'good': [x+1 for x in range(10)],
                     'leaning': [x+1 for x in range(10, 19)],
                     'javelin': [x+1 for x in range(20, 30)],
                     'align_arm': [x+1 for x in range(30, 40)],
                     'elbow_move': [x+1 for x in range(40, 50)]}

    # Algorithm(s) to use
    algos = {'k-means': {'n_clusters': 2}}

    clustering_problems = []

    for problem in datatype_joints_list:

        datatype_joints = problem[1]

        expert_sub_data = expert_data[:10] + expert_data[min(expert_data_repartion[problem[0]])-1:max(expert_data_repartion[problem[0]])]

        for algo, param in algos.items():
            print(problem[0])

            model = run_clustering(expert_sub_data, validate_data=False,
                                   datatype_joints=datatype_joints, algorithm=algo,
                                   parameters=param, scale_features=scale, normalise_features=normalise,
                                   true_labels=None, verbose=False, to_file=True, to_json=True, return_data=True)

        # Taking the student features
        std_features = data_gathering(student_data, datatype_joints)

        # If the expert data has been scaled, do the same for the student's ones
        if scale:
            from sklearn.preprocessing import RobustScaler
            scaler = RobustScaler()
            std_features = scaler.fit_transform(std_features)
        # Same for normalising
        if normalise:
            from sklearn.preprocessing import MinMaxScaler
            min_max_scaler = MinMaxScaler()
            std_features = min_max_scaler.fit_transform(std_features)

        # Compute the centroid of the student's features (Euclidean distance for now)
        std_centroid = get_centroid(std_features)

        # For the rest of the algorithm, if there are more than 2 features,
        # we run the data through a PCA for the next steps
        if len(model[0].cluster_centers_[0]) > 2:
            pca = PCA(n_components=2, copy=True)
            pca.fit(model[2])

            model[2] = pca.transform(model[2])
            model[0].cluster_centers_ = pca.transform(model[0].cluster_centers_)
            std_centroid = pca.transform(std_centroid.reshape(1, -1))[0]
            std_features = pca.transform(std_features)

        # Compute the distance from the student's centroid to the expert's ones
        distances_to_centroid = compute_distance(model[0].cluster_centers_, std_centroid)

        # Get the distance from the student's centroid to the line between the two expert's centroids
        distance_from_line = get_distance_from_expert_centoids_line(model[0].cluster_centers_, std_centroid)
        # Used to check if the student's centroid is between the expert centroids (diamond shape)
        distance_from_line /= (dst_pts(model[0].cluster_centers_[0], model[0].cluster_centers_[1]) / 2)

        # Normalise the distances to the centroids
        summ = sum(distances_to_centroid)
        for i, distance in enumerate(distances_to_centroid):
            distances_to_centroid[i] = distance/summ

        # Get the most probable cluster label for expert data
        clusters_label = get_cluster_label(expert_sub_data, expert_data_repartion, model[0].labels_)
        # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        #TODO: display numbers of motions in labelled clusters#
        # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        get_closeness_to_clusters(clusters_label, std_features, model[0].cluster_centers_, problem[0])

        # Display the closeness of the student's data to each expert cluster
        mix_res = mix(distances_to_centroid, clusters_label, distance_from_line)

        clusters_names = get_cluster_labels_from_data_repartition(model[0].labels_, model[0].cluster_centers_)

        clus_prob = ClusteringProblem(problem=problem[0],
                                      model=model[0], features=model[2],
                                      labels=model[0].labels_,
                                      centroids=model[0].cluster_centers_,
                                      sil_score=model[1]['ss'],
                                      clusters_names=clusters_names,
                                      algo_name=problem[0],
                                      std_data=std_features,
                                      std_centroid=get_centroid(std_features),
                                      distance_to_line=distance_from_line)

        clus_prob.trapezoid = get_trapezoid(model, std_centroid)
        clus_prob.circles = get_circle(model, std_centroid)
        clus_prob.std_data = std_features

        clustering_problems.append(clus_prob)

    give_two_advices(clustering_problems)
    plot_all_defaults(clustering_problems, only_centroids=False)

def only_feedback_new_descriptors(expert, student, path):
    folders_path = path
    if folders_path.split('/')[-1] == 'mixed':
        folders_path = "/".join(path.split('/')[:-1])

    # if not take_last_data(folders_path, student, expert, number=9):
    #     print(f'ERROR: {folders_path} does not exists')
    #     return

    # Expert Data
    expert_data = import_data(path, expert.name)
    # Student data
    student_data = import_data(path, student.name)

    # Setting the laterality
    for motion in expert_data:
        motion.laterality = expert.laterality

    for motion in student_data:
        motion.laterality = student.laterality

    # List of datatypes and joint to process
    # default to check: {Descriptor: [{joint, laterality or not (check left for lefthanded and vice versa)},
    #                                  other joint, laterality or not}],
    #                    Other descriptor: [{joint, laterality or not}]
    #                   }
    datatype_joints_list = []

    datatype_joints_list.append(['leaning', {'MeanSpeed': [{'joint': 'LeftShoulder', 'laterality': False},
                                                           {'joint': 'RightShoulder', 'laterality': False}]
                                }])

    datatype_joints_list.append(['elbow_move', {'MeanSpeed': [{'joint': 'LeftArm', 'laterality': True},
                                                              {'joint': 'LeftShoulder', 'laterality': True}]
                                }])

    datatype_joints_list.append(['javelin', {'DistanceX': [{'joint': 'distanceRightHandHead', 'laterality': True}],
                                             'DistanceY': [{'joint': 'distanceRightHandHead', 'laterality': True}],
                                             'DistanceZ': [{'joint': 'distanceRightHandHead', 'laterality': True}]
                                             }])

    datatype_joints_list.append(['align_arm', {'BoundingBoxWidthMean': [{'joint': 'HeadRightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxWidthStd': [{'joint': 'HeadRightShoulderRightArmRightForeArmRightHand', 'laterality': True}]
                                              }])



    # Scaling and normalisaing (nor not) the data
    # Useful for DBSCAN for example
    scale = False
    normalise = False

    expert_data_repartion = {'good': [x+1 for x in range(10)],
                     'leaning': [x+1 for x in range(10, 19)],
                     'javelin': [x+1 for x in range(20, 30)],
                     'align_arm': [x+1 for x in range(30, 40)],
                     'elbow_move': [x+1 for x in range(40, 50)]}

    # Algorithm(s) to use
    algos = {'k-means': {'n_clusters': 2}}

    clustering_problems = []

    for problem in datatype_joints_list:

        datatype_joints = problem[1]

        expert_sub_data = expert_data[:10] + expert_data[min(expert_data_repartion[problem[0]])-1:max(expert_data_repartion[problem[0]])]

        # if problem[0] == 'javelin':
            # del expert_sub_data[16]
            # del expert_sub_data[15]
            # del expert_sub_data[12]
            # del expert_sub_data[3]
            # del expert_sub_data[2]
            # del expert_sub_data[1]
            # del expert_sub_data[0]

        if problem[0] == 'align_arm':
            del expert_sub_data[15]
            del expert_sub_data[9]
            del expert_sub_data[8]


        for algo, param in algos.items():
            print(problem[0])

            model = run_clustering(expert_sub_data, validate_data=False,
                                   datatype_joints=datatype_joints, algorithm=algo,
                                   parameters=param, scale_features=scale, normalise_features=normalise,
                                   true_labels=None, verbose=False, to_file=True, to_json=True, return_data=True)

        # Taking the student features
        std_features = data_gathering(student_data, datatype_joints)

        # If the expert data has been scaled, do the same for the student's ones
        if scale:
            from sklearn.preprocessing import RobustScaler
            scaler = RobustScaler()
            std_features = scaler.fit_transform(std_features)
        # Same for normalising
        if normalise:
            from sklearn.preprocessing import MinMaxScaler
            min_max_scaler = MinMaxScaler()
            std_features = min_max_scaler.fit_transform(std_features)

        # Compute the centroid of the student's features (Euclidean distance for now)
        std_centroid = get_centroid(std_features)

        # Compute the distance from the student's centroid to the expert's ones
        distances_to_centroid = compute_distance(model[0].cluster_centers_, std_centroid)

        # Get the distance from the student's centroid to the line between the two expert's centroids
        distance_from_line = get_distance_from_expert_centoids_line(model[0].cluster_centers_, std_centroid)
        # Used to check if the student's centroid is between the expert centroids (diamond shape)
        distance_from_line /= (dst_pts(model[0].cluster_centers_[0], model[0].cluster_centers_[1]) / 2)

        # Normalise the distances to the centroids
        summ = sum(distances_to_centroid)
        for i, distance in enumerate(distances_to_centroid):
            distances_to_centroid[i] = distance/summ

        # Get the most probable cluster label for expert data
        clusters_label = get_cluster_label(expert_sub_data, expert_data_repartion, model[0].labels_)
        # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        #TODO: display numbers of motions in labelled clusters#
        # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        get_closeness_to_clusters(clusters_label, std_features, model[0].cluster_centers_, problem[0])

        # Display the closeness of the student's data to each expert cluster
        mix_res = mix(distances_to_centroid, clusters_label, distance_from_line)

        clusters_names = get_cluster_labels_from_data_repartition(model[0].labels_, model[0].cluster_centers_)

        # For the rest of the algorithm, if there are more than 2 features,
        # we run the data through a PCA for the next steps
        if len(model[0].cluster_centers_[0]) > 2:
            pca = PCA(n_components=2, copy=True)
            pca.fit(model[2])

            model[2] = pca.transform(model[2])
            model[0].cluster_centers_ = pca.transform(model[0].cluster_centers_)
            std_centroid = pca.transform(std_centroid.reshape(1, -1))[0]
            std_features = pca.transform(std_features)

        clus_prob = ClusteringProblem(problem=problem[0],
                                      model=model[0], features=model[2],
                                      labels=model[0].labels_,
                                      centroids=model[0].cluster_centers_,
                                      sil_score=model[1]['ss'],
                                      clusters_names=clusters_names,
                                      algo_name=problem[0],
                                      std_data=std_features,
                                      std_centroid=get_centroid(std_features),
                                      distance_to_line=distance_from_line)

        clus_prob.trapezoid = get_trapezoid(model, std_centroid)
        clus_prob.circles = get_circle(model, std_centroid)
        clus_prob.std_data = std_features

        clustering_problems.append(clus_prob)

    give_two_advices(clustering_problems)
    plot_all_defaults(clustering_problems, only_centroids=True)

    all_features_to_extract = merge_list_of_dictionnaries([x[1] for x in datatype_joints_list])
    expert_good_data = expert_data[min(expert_data_repartion['good'])-1:max(expert_data_repartion['good'])]
    plot_good_vs_student_all_data(expert_good_data, student_data, algos, all_features_to_extract, only_centroids=True)



def compute_distance_to_clusters(expert, student, path, begin, end):
    folders_path = path
    if folders_path.split('/')[-1] == 'mixed':
        folders_path = "/".join(path.split('/')[:-1])

    if not take_specific_data(folders_path, student, expert, begin=begin, end=end):
        print(f'ERROR: {folders_path} does not exists')
        return

    # Expert Data
    expert_data = import_data(path, expert.name)
    # Student data
    student_data = import_data(path, student.name)

    # Setting the laterality
    for motion in expert_data:
        motion.laterality = expert.laterality

    for motion in student_data:
        motion.laterality = student.laterality

    # List of datatypes and joint to process
    # default to check: {Descriptor: [{joint, laterality or not (check left for lefthanded and vice versa)},
    #                                  other joint, laterality or not}],
    #                    Other descriptor: [{joint, laterality or not}]
    #                   }
    datatype_joints_list = []

    datatype_joints_list.append(['leaning', {'MeanSpeed': [{'joint': 'LeftShoulder', 'laterality': False},
                                                           {'joint': 'RightShoulder', 'laterality': False}]
                                }])

    datatype_joints_list.append(['elbow_move', {'MeanSpeed': [{'joint': 'LeftArm', 'laterality': True},
                                                              {'joint': 'LeftShoulder', 'laterality': True}]
                                }])

    datatype_joints_list.append(['javelin', {'PosX': [{'joint': 'LeftHand', 'laterality': True},
                                                      {'joint': 'Head',     'laterality': False}],
                                             'PosY': [{'joint': 'LeftHand', 'laterality': True},
                                                      {'joint': 'Head',     'laterality': False}],
                                             'PosZ': [{'joint': 'LeftHand', 'laterality': True},
                                                      {'joint': 'Head',     'laterality': False}]
                                            }])

    datatype_joints_list.append(['align_arm', {'BoundingBoxMinusX': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxPlusX':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxMinusY': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxPlusY':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxMinusZ': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxPlusZ':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}]
                                              }])



    # Scaling and normalisaing (nor not) the data
    # Useful for DBSCAN for example
    scale = False
    normalise = False

    expert_data_repartion = {'good': [x+1 for x in range(10)],
                     'leaning': [x+1 for x in range(10, 19)],
                     'javelin': [x+1 for x in range(20, 30)],
                     'align_arm': [x+1 for x in range(30, 40)],
                     'elbow_move': [x+1 for x in range(40, 50)]}

    # Algorithm(s) to use
    algos = {'k-means': {'n_clusters': 2}}

    clustering_problems = []

    for problem in datatype_joints_list:

        datatype_joints = problem[1]

        expert_sub_data = expert_data[:10] + expert_data[min(expert_data_repartion[problem[0]])-1:max(expert_data_repartion[problem[0]])]

        for algo, param in algos.items():
            #print(problem[0])

            model = run_clustering(expert_sub_data, validate_data=False,
                                   datatype_joints=datatype_joints, algorithm=algo,
                                   parameters=param, scale_features=scale, normalise_features=normalise,
                                   true_labels=None, verbose=False, to_file=True, to_json=True, return_data=True)

        # Taking the student features
        std_features = data_gathering(student_data, datatype_joints)

        # If the expert data has been scaled, do the same for the student's ones
        if scale:
            from sklearn.preprocessing import RobustScaler
            scaler = RobustScaler()
            std_features = scaler.fit_transform(std_features)
        # Same for normalising
        if normalise:
            from sklearn.preprocessing import MinMaxScaler
            min_max_scaler = MinMaxScaler()
            std_features = min_max_scaler.fit_transform(std_features)

        # Compute the centroid of the student's features (Euclidean distance for now)
        std_centroid = get_centroid(std_features)

        # For the rest of the algorithm, if there are more than 2 features,
        # we run the data through a PCA for the next steps
        if len(model[0].cluster_centers_[0]) > 2:
            pca = PCA(n_components=2, copy=True)
            pca.fit(model[2])

            model[2] = pca.transform(model[2])
            model[0].cluster_centers_ = pca.transform(model[0].cluster_centers_)
            std_centroid = pca.transform(std_centroid.reshape(1, -1))[0]
            std_features = pca.transform(std_features)

        # Compute the distance from the student's centroid to the expert's ones
        distances_to_centroid = compute_distance(model[0].cluster_centers_, std_centroid)

        # Get the distance from the student's centroid to the line between the two expert's centroids
        distance_from_line = get_distance_from_expert_centoids_line(model[0].cluster_centers_, std_centroid)
        # Used to check if the student's centroid is between the expert centroids (diamond shape)
        distance_from_line /= (dst_pts(model[0].cluster_centers_[0], model[0].cluster_centers_[1]) / 2)

        # Normalise the distances to the centroids
        summ = sum(distances_to_centroid)
        for i, distance in enumerate(distances_to_centroid):
            distances_to_centroid[i] = distance/summ

        # Get the most probable cluster label for expert data
        clusters_label = get_cluster_label(expert_sub_data, expert_data_repartion, model[0].labels_)
        # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        #TODO: display numbers of motions in labelled clusters#
        # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        get_closeness_to_clusters(clusters_label, std_features, model[0].cluster_centers_, problem[0], name=student.name, to_print=False)

def rotated_feedback_comparison(expert, student, path, begin, end):
    folders_path = path
    if folders_path.split('/')[-1] == 'mixed':
        folders_path = "/".join(path.split('/')[:-1])

    if not take_specific_data(folders_path, student, expert, begin=begin, end=end):
        print(f'ERROR: {folders_path} does not exists')
        return

    # Expert Data
    expert_data = import_data(path, expert.name)
    # Student data
    student_data = import_data(path, student.name)

    # Setting the laterality
    for motion in expert_data:
        motion.laterality = expert.laterality

    for motion in student_data:
        motion.laterality = student.laterality

    # List of datatypes and joint to process
    # default to check: {Descriptor: [{joint, laterality or not (check left for lefthanded and vice versa)},
    #                                  other joint, laterality or not}],
    #                    Other descriptor: [{joint, laterality or not}]
    #                   }
    datatype_joints_list = []

    datatype_joints_list.append(['leaning', {'MeanSpeed': [{'joint': 'LeftShoulder', 'laterality': False},
                                                           {'joint': 'RightShoulder', 'laterality': False}]
                                }])

    datatype_joints_list.append(['elbow_move', {'MeanSpeed': [{'joint': 'LeftArm', 'laterality': True},
                                                              {'joint': 'LeftShoulder', 'laterality': True}]
                                }])

    datatype_joints_list.append(['javelin', {'Distances': [{'joint': 'distanceRightHandHead', 'laterality': True}]
                                }])

    datatype_joints_list.append(['align_arm', {'BoundingBoxWidthMean': [{'joint': 'HeadRightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                               'BoundingBoxWidthStd': [{'joint': 'HeadRightShoulderRightArmRightForeArmRightHand', 'laterality': True}]
                                              }])

    # datatype_joints_list.append(['javelin', {'PosX': [{'joint': 'LeftHand', 'laterality': True},
    #                                                   {'joint': 'Head',     'laterality': False}],
    #                                          'PosY': [{'joint': 'LeftHand', 'laterality': True},
    #                                                   {'joint': 'Head',     'laterality': False}],
    #                                          'PosZ': [{'joint': 'LeftHand', 'laterality': True},
    #                                                   {'joint': 'Head',     'laterality': False}]
    #                                         }])

    # datatype_joints_list.append(['align_arm', {'BoundingBoxMinusX': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
    #                                            'BoundingBoxPlusX':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
    #                                            'BoundingBoxMinusY': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
    #                                            'BoundingBoxPlusY':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
    #                                            'BoundingBoxMinusZ': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
    #                                            'BoundingBoxPlusZ':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}]
    #                                           }])



    # Scaling and normalisaing (nor not) the data
    # Useful for DBSCAN for example
    scale = False
    normalise = False

    expert_data_repartion = {'good': [x+1 for x in range(10)],
                     'leaning': [x+1 for x in range(10, 19)],
                     'javelin': [x+1 for x in range(20, 30)],
                     'align_arm': [x+1 for x in range(30, 40)],
                     'elbow_move': [x+1 for x in range(40, 50)]}

    # Algorithm(s) to use
    algos = {'k-means': {'n_clusters': 2}}

    clustering_problems = []

    for problem in datatype_joints_list:

        datatype_joints = problem[1]

        expert_sub_data = expert_data[:10] + expert_data[min(expert_data_repartion[problem[0]])-1:max(expert_data_repartion[problem[0]])]

        for algo, param in algos.items():
            print(problem[0])

            model = run_clustering(expert_sub_data, validate_data=False,
                                   datatype_joints=datatype_joints, algorithm=algo,
                                   parameters=param, scale_features=scale, normalise_features=normalise,
                                   true_labels=None, verbose=False, to_file=True, to_json=True, return_data=True)

        # Taking the student features
        std_features = data_gathering(student_data, datatype_joints)

        # If the expert data has been scaled, do the same for the student's ones
        if scale:
            from sklearn.preprocessing import RobustScaler
            scaler = RobustScaler()
            std_features = scaler.fit_transform(std_features)
        # Same for normalising
        if normalise:
            from sklearn.preprocessing import MinMaxScaler
            min_max_scaler = MinMaxScaler()
            std_features = min_max_scaler.fit_transform(std_features)

        # Compute the centroid of the student's features (Euclidean distance for now)
        std_centroid = get_centroid(std_features)

        # For the rest of the algorithm, if there are more than 2 features,
        # we run the data through a PCA for the next steps
        if len(model[0].cluster_centers_[0]) > 2:
            pca = PCA(n_components=2, copy=True)
            pca.fit(model[2])

            model[2] = pca.transform(model[2])
            model[0].cluster_centers_ = pca.transform(model[0].cluster_centers_)
            std_centroid = pca.transform(std_centroid.reshape(1, -1))[0]
            std_features = pca.transform(std_features)

        # Compute the distance from the student's centroid to the expert's ones
        distances_to_centroid = compute_distance(model[0].cluster_centers_, std_centroid)

        # Get the distance from the student's centroid to the line between the two expert's centroids
        distance_from_line = get_distance_from_expert_centoids_line(model[0].cluster_centers_, std_centroid)
        # Used to check if the student's centroid is between the expert centroids (diamond shape)
        distance_from_line /= (dst_pts(model[0].cluster_centers_[0], model[0].cluster_centers_[1]) / 2)

        # Normalise the distances to the centroids
        summ = sum(distances_to_centroid)
        for i, distance in enumerate(distances_to_centroid):
            distances_to_centroid[i] = distance/summ

        # Get the most probable cluster label for expert data
        clusters_label = get_cluster_label(expert_sub_data, expert_data_repartion, model[0].labels_)
        # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        #TODO: display numbers of motions in labelled clusters#
        # # # # # # # # # # # # # # # # # # # # # # # # # # # #
        get_closeness_to_clusters(clusters_label, std_features, model[0].cluster_centers_, problem[0])

        # Display the closeness of the student's data to each expert cluster
        mix_res = mix(distances_to_centroid, clusters_label, distance_from_line)

        clusters_names = get_cluster_labels_from_data_repartition(model[0].labels_, model[0].cluster_centers_)

        clus_prob = ClusteringProblem(problem=problem[0],
                                      model=model[0], features=model[2],
                                      labels=model[0].labels_,
                                      centroids=model[0].cluster_centers_,
                                      sil_score=model[1]['ss'],
                                      clusters_names=clusters_names,
                                      algo_name=problem[0],
                                      std_data=std_features,
                                      std_centroid=get_centroid(std_features),
                                      distance_to_line=distance_from_line)

        clus_prob.trapezoid = get_trapezoid(model, std_centroid)
        clus_prob.circles = get_circle(model, std_centroid)
        clus_prob.std_data = std_features

        clustering_problems.append(clus_prob)

    return give_two_advices(clustering_problems)



def take_last_data(path, student, expert, number=9):

    if not student.full_name in os.listdir(path):
        print(f'ERROR: {student.full_name} not found in {path}.')
        return False

    mixed_path = os.path.normpath(os.path.join(path, 'mixed'))

    # Remove all non-expert data )
    for file in os.listdir(mixed_path):
        if expert.name not in file:
            print(f'Removed {file}')
            rmtree(os.path.normpath(os.path.join(mixed_path, file)))

    std_path = os.path.normpath(os.path.join(path, student.full_name))

    file_list = os.listdir(std_path)
    file_list = sorted(file_list, key=lambda sin: int(sin.replace(student.name + '_', '').replace('Char00', '')))
    folders_to_copy = file_list[-number:]

    for folder_to_copy in folders_to_copy:
        std_folder = os.path.join(std_path, folder_to_copy)
        print(f'Copying {folder_to_copy}')
        copy_tree(std_folder, os.path.join(mixed_path, folder_to_copy))

    return True

def take_specific_data(path, student, expert, begin=0, end=9, verbose=False):

    if not student.full_name in os.listdir(path):
        print(f'ERROR: {student.full_name} not found in {path}.')
        return False

    mixed_path = os.path.normpath(os.path.join(path, 'mixed'))

    # Remove all non-expert data
    for file in os.listdir(mixed_path):
        if expert.name not in file:
            if verbose:
                print(f'Removed {file}')
            rmtree(os.path.normpath(os.path.join(mixed_path, file)))

    std_path = os.path.normpath(os.path.join(path, student.full_name))

    file_list = os.listdir(std_path)
    file_list = sorted(file_list, key=lambda sin: int(sin.replace(student.name + '_', '').replace('Char00', '')))

    folders_to_copy = file_list[begin:end]

    for folder_to_copy in folders_to_copy:
        std_folder = os.path.join(std_path, folder_to_copy)
        if verbose:
            print(f'Copying {folder_to_copy}')
        copy_tree(std_folder, os.path.join(mixed_path, folder_to_copy))

    return True

def test_student_list():

    expert = Person(r'', 'aurel', 'Right')
    path = r'C:/Users/quentin/Documents/Programmation/C++/MLA/Data/alldartsdescriptors/students/mixed'

    for student in cst.students_list:
        print(f'Testing {student.name}')

        folders_path = path
        if folders_path.split('/')[-1] == 'mixed':
            folders_path = "/".join(path.split('/')[:-1])

        if not take_specific_data(folders_path, student, expert, begin=0, end=1):
            print(f'ERROR AT {student.name}')
            return

        # Expert Data
        expert_data = import_data(path, expert.name)
        # Student data
        student_data = import_data(path, student.name)

        # Setting the laterality
        for motion in expert_data:
            motion.laterality = expert.laterality

        for motion in student_data:
            motion.laterality = student.laterality

        datatype_joints_list = []

        datatype_joints_list.append(['leaning', {'MeanSpeed': [{'joint': 'LeftShoulder', 'laterality': False},
                                                               {'joint': 'RightShoulder', 'laterality': False}]
                                    }])

        datatype_joints_list.append(['elbow_move', {'MeanSpeed': [{'joint': 'LeftArm', 'laterality': True},
                                                                  {'joint': 'LeftShoulder', 'laterality': True}]
                                    }])

        datatype_joints_list.append(['javelin', {'PosX': [{'joint': 'LeftHand', 'laterality': True},
                                                          {'joint': 'Head',     'laterality': False}],
                                                 'PosY': [{'joint': 'LeftHand', 'laterality': True},
                                                          {'joint': 'Head',     'laterality': False}],
                                                 'PosZ': [{'joint': 'LeftHand', 'laterality': True},
                                                          {'joint': 'Head',     'laterality': False}]
                                                }])

        datatype_joints_list.append(['align_arm', {'BoundingBoxMinusX': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                                   'BoundingBoxPlusX':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                                   'BoundingBoxMinusY': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                                   'BoundingBoxPlusY':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                                   'BoundingBoxMinusZ': [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}],
                                                   'BoundingBoxPlusZ':  [{'joint': 'RightShoulderRightArmRightForeArmRightHand', 'laterality': True}]
                                                  }])



        # Scaling and normalisaing (nor not) the data
        # Useful for DBSCAN for example
        scale = False
        normalise = False

        expert_data_repartion = {'good': [x+1 for x in range(10)],
                         'leaning': [x+1 for x in range(10, 19)],
                         'javelin': [x+1 for x in range(20, 30)],
                         'align_arm': [x+1 for x in range(30, 40)],
                         'elbow_move': [x+1 for x in range(40, 50)]}

        # Algorithm(s) to use
        algos = {'k-means': {'n_clusters': 2}}

        clustering_problems = []

        for problem in datatype_joints_list:
            std_features = data_gathering(student_data, problem[1])

def export_advices_to_xlsx(path):
    from xlsxwriter import Workbook

    expert = Person(r'', 'aurel', 'Right')

    workbook = Workbook('new_advices_after_rotation.xlsx')

    for student in cst.students_list:
        print(f'Processing {student.name}...')
        worksheet = workbook.add_worksheet(student.name)
        worksheet.merge_range('A1:B1', 'New')
        worksheet.merge_range('C1:D1', 'Old')
        for i in range(4):
            advices = rotated_feedback_comparison(expert, student, path, begin=i*9, end=(i*9)+9)
            worksheet.write(f'A{i+2}', advices[0])
            worksheet.write(f'B{i+2}', advices[1])

        print('Done.')

    workbook.close()

    # # test_student_list()

def merge_xlsx():
    from openpyxl import load_workbook
    wb_results = load_workbook('csv_test/results_2.xlsx')
    wb_comparison = load_workbook('new_advices_after_rotation.xlsx')

    for student in cst.students_list:
        print(f'Processing {student.name}')
        try:
            sheet_results = wb_results[student.name[:-1]]
        except KeyError:
            sheet_results = wb_results['Saupin']
        sheet_comparison = wb_comparison[student.name]

        for i in range(3):
            result_string = sheet_results[f'E{((i+1)*10 + 3)}'].value
            result_string = result_string.split('/')
            sheet_comparison[f'C{i+2}'].value = result_string[0]
            sheet_comparison[f'D{i+2}'].value = result_string[1]

    wb_comparison.save('new_advices_comparison_one_sheet.xlsx')

def one_sheet_xlsx():
    from openpyxl import Workbook, load_workbook
    wb_results = load_workbook('csv_test/results_2.xlsx')
    wb_comparison = load_workbook('new_advices_after_rotation.xlsx')
    wb_final = Workbook()
    final_sheet = wb_final.active

    final_sheet[f'A1'].value = 'Nom'
    final_sheet[f'B1'].value = 'Old'
    final_sheet[f'D1'].value = 'New'


    for row, student in enumerate(cst.students_list):
        good = "True"

        # writting offset
        wo = 3 * row

        print(f'Processing {student.name}')
        try:
            sheet_results = wb_results[student.name[:-1]]
        except KeyError:
            sheet_results = wb_results['Saupin']
        sheet_comparison = wb_comparison[student.name]


        final_sheet[f'A{1 + 1 + wo}'].value = student.name

        for i in range(3):
            result_string = sheet_results[f'E{((i+1)*10 + 3)}'].value
            result_string = result_string.split('/')
            comparison_string = sheet_comparison[f'A{i+2}:B{i+2}']
            comparison_string = [sheet_comparison[f'A{i+2}'].value, sheet_comparison[f'B{i+2}'].value]

            result_string[0] = result_string[0].replace("_", " ").replace("Me\xa0: ", "").lower().strip()
            result_string[1] = result_string[1].replace("_", " ").replace("Me\xa0: ", "").lower().strip()
            comparison_string[0] = comparison_string[0].replace("_", " ").replace("Me : ", "").lower().strip()
            comparison_string[1] = comparison_string[1].replace("_", " ").replace("Me : ", "").lower().strip()

            final_sheet[f'B{i + 1 + 1 + wo}'].value = result_string[0]
            final_sheet[f'C{i + 1 + 1 + wo}'].value = result_string[1]
            final_sheet[f'D{i + 1 + 1 + wo}'].value = comparison_string[0]
            final_sheet[f'E{i + 1 + 1 + wo}'].value = comparison_string[1]

            if result_string[0] != comparison_string[0] or result_string[1] != comparison_string[1]:
                good = "False"

        final_sheet[f'G{wo + 3}'].value = good


    wb_final.save('blblbl.xlsx')

if __name__ == '__main__':
    expert = Person(r'', 'aurel', 'Right')
    student = Person(r'', 'JonardA', 'Right', 'Jonard_Antoine')
    path = r'C:/Users/quentin/Documents/Programmation/C++/MLA/Data/alldartsdescriptors/students/mixed'
    path = r'C:/Users/quentin/Documents/Programmation/C++/MLA/Data/alldartsdescriptors/test/noneed_rotated/mixed'
    # path = r'C:/Users/quentin/Documents/Programmation/C++/MLA/Data/alldartsdescriptors/clarete/mixed'
    # only_feedback(expert, student, path)
    only_feedback_new_descriptors(expert, student, path)

    # export_advices_to_xlsx(path)
    # one_sheet_xlsx()
