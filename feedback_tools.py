import os
from math import floor, sqrt
import operator
import csv

from dataclasses import dataclass

import numpy as np

from scipy.spatial import distance

import matplotlib.path as mplPath

from constants import problemes_et_solutions as problems_and_advices

from sklearn.decomposition import PCA

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

@dataclass
class Circle:
    def __init__(self, center, radius, limits=None, is_good=False):
        self.center = center
        self.radius = radius
        self.limits = limits
        # TODO: add limits check
        self.is_good = is_good

@dataclass
class Trapezoid:
    def __init__(self, p1, p2, p3, p4, path=None):
        self.p1 = p1
        self.p2 = p2
        self.p3 = p3
        self.p4 = p4
        self.path = path
        if not self.path:
            self.path = mplPath.Path(np.array([p1, p2, p3, p4]))

    def get_max_distance(self):
        return max([np.linalg.norm(x - y) for x in [np.asarray(self.p1), np.asarray(self.p2), np.asarray(self.p3), np.asarray(self.p4)]
                                          for y in [np.asarray(self.p1), np.asarray(self.p2), np.asarray(self.p3), np.asarray(self.p4)]])

@dataclass
class ClusteringProblem:
    def __init__(self, problem=None, model=None, features=None,
                 centroids=None, labels=None, sil_score=None,
                 algo_name=None, clusters_names=None, std_data=None,
                 std_centroid=None, trapezoid=None, circles=None,
                 distance_to_line=None):
        self.problem = problem
        self.model = model
        self.features = features
        self.centroids = centroids
        self.labels = labels
        self.sil_score = sil_score
        self.algo_name = algo_name
        self.clusters_names = clusters_names

        self.std_data = std_data
        self.std_centroid = std_centroid

        self.trapezoid = trapezoid
        self.circles = circles

        self.distance_to_line = distance_to_line

def intra_cluster_dst(centroid, points):
    return np.mean(np.asarray([np.linalg.norm(point - centroid) for point in points]))

def compute_distance(centroids, feature):
    return [np.linalg.norm(feature - centroid) for centroid in centroids]

def compute_single_distance(pt1, pt2):
    return np.linalg.norm(pt1 - pt2)

def dst_pts(pt1, pt2):
    if len(pt1) == 1:
        return abs(pt2 - pt1)[0]

    return sqrt(pow(pt2[0] - pt1[0], 2) + pow(pt2[1] - pt1[1], 2))


def get_trapezoid(result, point):
    centroids = result[0].cluster_centers_
    points = [result[2][np.where(result[0].labels_ == i)] for i in np.unique(result[0].labels_)]
    max_dst = [max(max(distance.cdist([centroids[i]], cluster, 'euclidean'))) for i, cluster in enumerate(points)]

    if len(centroids[0]) == 1:
        p1 = (centroids[0][0], 0 + max_dst[0])
        p4 = (centroids[0][0], 0 - max_dst[0])
        p2 = (centroids[1][0], 0 + max_dst[1])
        p3 = (centroids[1][0], 0 - max_dst[1])
    else:
        p1 = (centroids[0][0], centroids[0][1] + max_dst[0])
        p4 = (centroids[0][0], centroids[0][1] - max_dst[0])
        p2 = (centroids[1][0], centroids[1][1] + max_dst[1])
        p3 = (centroids[1][0], centroids[1][1] - max_dst[1])

    trapezoid = Trapezoid(p1, p2, p3, p4)
    # trapezoid = mplPath.Path(np.array([p1, p2, p3, p4]))
    # print(f"Point is in trapezoid: {is_in_trapezoid(point, trapezoid.path)}")
    return trapezoid

def is_in_trapezoid(point, trapezoid):
    point_to_check = point
    if len(point_to_check) == 1:
        point_to_check = np.array([point, 0.0])

    return trapezoid.contains_point(point_to_check)


def get_circle(result, point):
    centroids = result[0].cluster_centers_
    points = [result[2][np.where(result[0].labels_ == i)] for i in np.unique(result[0].labels_)]

    if len(centroids[0]) == 1:
        centroids = np.asarray([np.array([x[0], 0.0]) for x in result[0].cluster_centers_])
        points_tmp = []
        for subpoints in points:
            points_tmp.append(np.asarray([np.array([x[0], 0.0]) for x in subpoints]))
        points = points_tmp


    good_or_bad = np.ndarray.tolist(np.where(result[0].labels_ == 0)[0])

    is_good = False
    if len(set(good_or_bad).intersection([x for x in range(10)])) > 5:
        is_good = True

    max_dst = [max(max(distance.cdist([centroids[i]], cluster, 'euclidean'))) for i, cluster in enumerate(points)]

    med_dst = [intra_cluster_dst(centroid, points[i]) for i, centroid in enumerate(centroids)]

    is_in_cluster = is_in_circle(point, centroids[0], max_dst[0])
    if not is_in_cluster:
        is_in_cluster = is_in_circle(point, centroids[1], max_dst[1])

    circle_good = Circle(centroids[0], max_dst[0],
                         limits={'center': centroids[0],
                                 'radius_max': max_dst[0],
                                 'radius_med': (max_dst[0] + med_dst[0]) / 2,
                                 'radius_min': med_dst[0]},
                         is_good=is_good)

    circle_bad = Circle(centroids[1], max_dst[1],
                        limits={'center': centroids[1],
                                'radius_max': max_dst[1],
                                'radius_med': (max_dst[1] + med_dst[1]) / 2,
                                'radius_min': med_dst[1]},
                        is_good=not is_good)

    return (circle_good, circle_bad)

def is_in_circle(point, centroid, diameter):
    return np.linalg.norm(point - centroid) < diameter

def is_in_circle_c(point, circle):
    return np.linalg.norm(point - circle.center) < circle.radius

def get_cluster_label(original_data, original_labels, clustering_labels):

    labelled_cluster = {}
    labelled_data = []
    for data in original_data:
        # Get rid of the name + the Char00
        # Since I changed the name of the files, we catch the ValueError
        # exception if it's the old format
        # new : [name]_[number]Char00
        # old : [name]_[number]_Char00
        try:
            number = int(data.name.split("_")[-2])
        except ValueError:
            number = int(data.name.split("_")[1].replace("Char00", ""))

        for k, v in original_labels.items():
            if number in v:
                number = k
                break

        if not isinstance(number, int):
            labelled_data.append(number)

    c_rep = []
    for cluster in set(clustering_labels):
        c_rep.append({x: 0 for x in original_labels.keys()})
        for i, elem in enumerate(clustering_labels):
            if elem == cluster:
                c_rep[-1][labelled_data[i]] += 1


    for i, cluster in enumerate(c_rep):
        labelled_cluster[f'c{i}'] = max(cluster.items(), key=operator.itemgetter(1))[0]

    return labelled_cluster

def get_centroid(data):
    return np.mean(data, axis=0)

def get_distance_from_expert_centoids_line(exp_centroids, std_centroid):
    if len(exp_centroids[0]) == 1:
        return 0

    pt1 = {'x': exp_centroids[0][0], 'y': exp_centroids[0][1]}
    pt2 = {'x': exp_centroids[1][0], 'y': exp_centroids[1][1]}

    pt_std = {'x': std_centroid[0], 'y': std_centroid[1]}


    A = pt2['y'] - pt1['y']
    B = pt2['x'] - pt1['x']
    C = (pt1['x'] * pt2['y']) - (pt2['x'] * pt1['y'])

    return abs((A * pt_std['x']) + (B * pt_std['y']) + C) / sqrt(pow(A, 2) + pow(B, 2))

def mix(distances, c_labels, distance_line_vs_distance_centroids):
    new_dict = {}
    for i, (_, v) in enumerate(c_labels.items()):
        new_dict[v] = distances[i]


    keys_to_display = list(new_dict.keys())
    if keys_to_display[0] != 'good':
        keys_to_display.reverse()

    size_display = 40
    char = '#'
    place = floor((size_display / 100) * (100 * new_dict['good']))

    print(f"{keys_to_display[0]} [", end='')
    for i in range(size_display):
        if i == place:
            print(char, end='')
        else:
            print('-', end='')
    print(f"] {keys_to_display[1]} ({distance_line_vs_distance_centroids:.2f})")

    return (keys_to_display, distance_line_vs_distance_centroids)

def get_closeness_to_clusters(c_labels, features, cps, problem, name='', to_print=True, to_csv=True):
    # TODO: rename the variables to something at least barely explicit ffs
    new_dict = {}
    label_lst = []
    for i, (_, v) in enumerate(c_labels.items()):
        new_dict[v] = []
        label_lst.append(v)

    lst = []
    for i, feat in enumerate(features):
        distance = compute_distance(cps, feat)
        lst.append({"sample": i, label_lst[0]: round(distance[0], 4), label_lst[1]: round(distance[1], 4)})

    if to_print:
        for sample in lst:
            # print(f"{sample['sample']} {sample['good']} {sample[problem]} {problem}")
            print(f"{sample['good']} {sample[problem]}")

    if to_csv:
        path_to_file = f'{problem}_output.csv'
        if name:
            path_to_file = f'{name}_' + path_to_file
        path_to_file = 'csv_test/' + path_to_file

        with open(path_to_file, 'a', newline='') as out_csv:
            data_distance_writer = csv.writer(out_csv, delimiter=',', quotechar='"')
            data_distance_writer.writerow('')
            for sample in lst:
                data_distance_writer.writerow([sample['good'], sample[problem]])

def get_global_closeness_to_good_cluster(cps, removed_values, name='', to_print=True, to_csv=True):
    # Take all features from 4 problems
    # normalise by mean for each dimension for each problem
    # merge features vectors
    # compute distance

    # HUGE PROBLEM : due to the clustering, good and bad clusters don't have
    # the same amount of expert data in each from one problem to another...
    # Features Normalisation
    normalised_features = {}

    removed_values_set = set()
    for key in removed_values:
        removed_values_set.update(removed_values[key])

    removed_values_set = sorted(removed_values_set, reverse=True)

    # for each problem
    for cp in cps:
        # Concatenate expert and student features, while deleting the expert
        # outliers

        all_features = np.concatenate((np.delete(cp.features, removed_values_set, axis=0), cp.std_data))
        exp_features_number = len(cp.features)

        # Divide each feature by the mean of each axis
        all_features /= np.mean(all_features, axis=0)

        # Put them into a dict (keys: expert, student)
        normalised_features[cp.problem] = {}

        labels_to_use = cp.labels.copy()
        #TODO: not ad-hoc
        if cp.problem == 'align_arm':
            labels_to_use = np.delete(labels_to_use, -1)

        if cp.problem == 'leaning':
            labels_to_use = np.delete(labels_to_use, [7, 8, 14])

        else:
            labels_to_use = np.delete(labels_to_use, [8, 9, 15, 19])

        good_idx = [x for x, val in enumerate(cp.labels) if val == 1]
        bad_idx = [x for x, val in enumerate(cp.labels) if val == 0]

        # Swap them (because I know that idx from 0 to 10 are the good ones)
        if good_idx[0] > bad_idx[0]:
            good_idx, bad_idx = bad_idx, good_idx

        normalised_features[cp.problem]['expert_good'] = all_features[good_idx]
        normalised_features[cp.problem]['expert_bad'] = all_features[bad_idx]
        normalised_features[cp.problem]['student'] = all_features[exp_features_number:]

    merged_vectors = {'expert_good': None, 'expert_bad': None, 'student': None}

    # 3 outputs : expert good values / bad values, std values
    for key in merged_vectors.keys():
        merged_features = []

        for problem in normalised_features.keys():
            if not isinstance(merged_features, np.ndarray):
                merged_features = normalised_features[problem][key].copy()
            else:
                try:
                    np.concatenate((merged_features, normalised_features[problem][key].copy()), axis=1)
                except:
                    breakpoint()

        breakpoint()

def get_global_closeness_to_good_cluster_2(cps, name='', to_print=True, to_csv=True, to_xlsx=False):
    # Take all features from 4 problems
    # normalise by mean for each dimension for each problem
    # merge features vectors
    # compute distance

    # Features Normalisation
    normalised_features = {}

    # for each problem
    for cp in cps:
        # Concatenate expert and student features, while deleting the expert
        # outliers
        all_features = np.concatenate((cp.features[:8], cp.std_data))
        exp_features_number = 8

        # Divide each feature by the mean of each axis
        all_features /= np.mean(all_features, axis=0)

        # Put them into a dict (keys: expert, student)
        normalised_features[cp.problem] = {}

        normalised_features[cp.problem]['expert_good'] = all_features[:exp_features_number]
        normalised_features[cp.problem]['student'] = all_features[exp_features_number:]

    merged_vectors = {'expert_good': None, 'student': None}


    # 3 outputs : expert good values / bad values, std values
    for key in merged_vectors.keys():
        merged_features = []

        for problem in normalised_features.keys():
            if not isinstance(merged_features, np.ndarray):
                merged_features = normalised_features[problem][key].copy()
            else:
                merged_features = np.concatenate((merged_features, normalised_features[problem][key].copy()), axis=1)

        merged_vectors[key] = merged_features.copy()

    final_components = len(merged_vectors['expert_good'][0])
    dst = {}

    for i in range(2, final_components + 1):
        pca = PCA(n_components=i, copy=True)
        full_data = np.concatenate((merged_vectors['expert_good'], merged_vectors['student']))
        pca.fit(full_data)

        merged_vectors_2 = {}

        merged_vectors_2['expert_good'] = pca.transform(merged_vectors['expert_good'])
        merged_vectors_2['student'] = pca.transform(merged_vectors['student'])

        expert_centroid = get_centroid(merged_vectors_2['expert_good'])

        dst[i] = [compute_single_distance(expert_centroid, x) for x in merged_vectors_2['student']]


    if to_print:
        pass

    if to_csv:
        path_to_file = f'dst_all_pca_2_{final_components}_output.csv'
        if name:
            path_to_file = f'{name}_' + path_to_file
        path_to_file = 'csv_test/' + path_to_file

        with open(path_to_file, 'a', newline='') as out_csv:
            data_distance_writer = csv.writer(out_csv, delimiter=',', quotechar='"')
            data_distance_writer.writerow(['Distance to good'])
            data_distance_writer.writerow(['PCA Value'])
            data_distance_writer.writerow(list(dst.keys()))
            transposed_data = list(zip(*dst.values()))
            for i, row in enumerate(transposed_data):
                # csv_writer needs a list to write
                if i > 0 and i % 9 == 0:
                    data_distance_writer.writerow([''])
                data_distance_writer.writerow(row)

    if to_xlsx:
        wb = load_workbook(filename='csv_test/distance_pca/results_3.xlsx')
        ws = wb[name]

        ws['O1'].value = 'All merged, distance to expert\'s good cluster'
        font_style = Font(size='14')
        cell_alignment = Alignment(horizontal='center')
        ws['O1'].font = font_style
        ws['O1'].alignment = cell_alignment
        ws.merge_cells('O1:V1')
        ws['O2'].value = 'PCA Value'
        ws['O2'].font = font_style
        ws['O2'].alignment = cell_alignment
        ws.merge_cells('O2:V2')
        for i, key in enumerate(dst.keys()):
            ws[f'{chr(79+i)}3'].font = font_style
            ws[f'{chr(79+i)}3'].alignment = cell_alignment
            ws[f'{chr(79+i)}3'].value = key

        row_idx = 3
        transposed_data = list(zip(*dst.values()))

        for i, row in enumerate(transposed_data):

            row_idx += 1
            if i > 0 and i % 9 == 0:
                row_idx += 1

            for j, val in enumerate(row):
                ws[f'{chr(79+j)}{row_idx}'].value = val

        wb.save('csv_test/distance_pca/results_3.xlsx')



def csv_merger(path, name=None):
    # csv_align_arm = open(path + f'{name}_align_arm_output.csv')
    # csv_elbow_move = open(path + f'{name}_elbow_move_output.csv')
    # csv_javelin = open(path + f'{name}_javelin_output.csv')
    # csv_leaning = open(path + f'{name}_leaning_output.csv')

    files_to_concatenate = [f'{path}{name}_leaning_output.csv',
                            f'{path}{name}_elbow_move_output.csv',
                            f'{path}{name}_javelin_output.csv',
                            f'{path}{name}_align_arm_output.csv']

    temp_data = []
    for filenames in files_to_concatenate:
        temp_data.append(np.loadtxt(filenames, dtype='str'))

    temp_data = np.array(temp_data)
    np.savetxt(f'{path}{name}_merged.csv', temp_data.transpose(), fmt='%s', delimiter=',')

    # with open(f'{path}_{name}_merged.csv', 'w') as final_csv_out:


    # csv_align_arm.close()
    # csv_elbow_move.close()
    # csv_javelin.close()
    # csv_leaning.close()

def clean_csv_folder(path, verbose=False):
    for file in os.listdir(path):
        if '.csv' in file and 'merged' not in file:
            if verbose:
                print(f'Removed {file}')
            os.remove(os.path.normpath(os.path.join(path, file)))

def get_cluster_labels_from_data_repartition(labels, centroids):
    good_or_bad = np.ndarray.tolist(np.where(labels == 0)[0])

    if len(set(good_or_bad).intersection([x for x in range(10)])) > 5:
        return('good', 'bad')

    return ('bad', 'good')

def give_advice(clustering_problems):
    candidates = [x for x in clustering_problems if is_in_trapezoid(x.std_centroid, x.trapezoid.path)]
    if candidates:
        # dst(std, good) / dst(good, bad) -> normalisation
        dst_to_closest = [np.linalg.norm((x.std_centroid - x.centroids[x.clusters_names.index('good')]) / x.trapezoid.get_max_distance()) for x in candidates]
        print(dst_to_closest)
        advice_to_give = candidates[dst_to_closest.index(min(dst_to_closest))].problem
        print(f"Problem: {advice_to_give}")
        print(problems_and_advices[advice_to_give])

def get_indexes(candidates, second_pass=False, biggest=False):
    first_dst = None
    second_dst = None

    all_dst = []

    if len(candidates) > 0:
        all_dst = [np.linalg.norm((x.std_centroid - x.centroids[x.clusters_names.index('good')]) / x.trapezoid.get_max_distance()) for x in candidates]
        first_dst = all_dst.index(min(all_dst))
        if second_pass:
            if biggest:
                return all_dst.index(max(all_dst))
            return first_dst

    if len(candidates) > 1:
        all_dst[first_dst] = max(all_dst) + 1
        second_dst = all_dst.index(min(all_dst))

    return first_dst, second_dst

def give_two_advices(clustering_problems):

    first_advice_to_give = None
    second_advice_to_give = None
    # Candidates are (In the trapezoid OR in the bad cluster) AND NOT in the good cluster
    candidates = [x for x in clustering_problems if (is_in_trapezoid(x.std_centroid, x.trapezoid.path)
                                                     or is_in_circle_c(x.std_centroid, next((y for y in x.circles if y.is_good == False), None)))
                                                 and not is_in_circle_c(x.std_centroid, next((y for y in x.circles if y.is_good == True), None))]

    if len(candidates) > 1:
        print(f"2 or more candidates ({len(candidates)})")

        # dst(std, good) / dst(good, bad) -> normalisation
        first_advice_idx, second_advice_idx = get_indexes(candidates)

        first_advice_to_give = candidates[first_advice_idx].problem
        print(f"Problem: {first_advice_to_give}\n {problems_and_advices[first_advice_to_give]}")
        second_advice_to_give = candidates[second_advice_idx].problem
        print(f"Problem: {second_advice_to_give}\n {problems_and_advices[second_advice_to_give]}")

    elif len(candidates) == 1:
        print("1 candidate")

        # dst(std, good) / dst(good, bad) -> normalisation
        first_advice_idx, _ = get_indexes(candidates)
        first_advice_to_give = candidates[first_advice_idx].problem
        print(f"Problem: {first_advice_to_give}\n {problems_and_advices[first_advice_to_give]}")

        new_candidates = [x for x in clustering_problems if not is_in_circle_c(x.std_centroid, next((y for y in x.circles if y.is_good == True), None))
                                                         and x.problem != first_advice_to_give]
        biggest = False
        # If everything is in good clusters (lul)
        if not new_candidates:
            new_candidates = [x for x in clustering_problems if x.problem != first_advice_to_give]
            biggest = True

        second_advice_idx = get_indexes(new_candidates, second_pass=True, biggest=biggest)

        second_advice_to_give = new_candidates[second_advice_idx].problem
        print(f"Problem: {second_advice_to_give}\n {problems_and_advices[second_advice_to_give]}")

    else:
        print("No candidate")

        new_candidates = [x for x in clustering_problems if not is_in_circle_c(x.std_centroid, next((y for y in x.circles if y.is_good == True), None))]

        max_val = max([x.distance_to_line for x in clustering_problems])
        dst_to_line = []
        for x in clustering_problems:
            if (not is_in_trapezoid(x.std_centroid, x.trapezoid.path)
                and not is_in_circle_c(x.std_centroid, next((y for y in x.circles if y.is_good == True), None))):
                dst_to_line.append(x.distance_to_line)
            else:
                dst_to_line.append(max_val + 1)


        first_advice = dst_to_line.index(min(dst_to_line))
        first_advice_to_give = clustering_problems[first_advice].problem
        print(f"Problem: {first_advice_to_give}\n {problems_and_advices[first_advice_to_give]}")

        dst_to_line[first_advice] = max(dst_to_line) + 1
        second_advice = dst_to_line.index(min(dst_to_line))
        second_advice_to_give = clustering_problems[second_advice].problem
        print(f"Problem: {second_advice_to_give}\n {problems_and_advices[second_advice_to_give]}")

    return (first_advice_to_give, second_advice_to_give)
